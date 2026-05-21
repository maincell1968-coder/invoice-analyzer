import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import io

# Key column names matching app.py
COL_SPEDIZIONE = "Numero spedizione principale"
COL_DESCRIZIONE = "Descrizione spesa"


def generate_excel_report(df, grouped, rules, client_code):
    """
    Generates a beautifully styled corporate Excel report for the client
    summarizing costs, inefficiencies, and actionable recommendations.
    
    Returns a bytes object of the generated Excel workbook.
    """
    wb = openpyxl.Workbook()
    
    # Define Palette (UPS Corporate Premium Palette: Chocolate Brown, Gold Accent, Cream Background)
    color_primary = "36220F"      # UPS Dark Chocolate Brown
    color_primary_light = "F4F0EA" # Cream White
    color_accent = "D4AF37"       # Gold
    color_accent_light = "FFFCEB" # Light Gold
    color_text_light = "FFFFFF"   # White
    color_alert = "FADBD8"        # Soft Red
    color_warning = "FCF3CF"      # Soft Yellow
    color_ok = "D5F5E3"           # Soft Green
    color_gray_border = "D5D8DC"  # Light Gray
    
    # Styles
    font_title = Font(name="Calibri", size=16, bold=True, color=color_text_light)
    font_section = Font(name="Calibri", size=13, bold=True, color=color_primary)
    font_header = Font(name="Calibri", size=11, bold=True, color=color_text_light)
    font_bold = Font(name="Calibri", size=11, bold=True, color=color_primary)
    font_regular = Font(name="Calibri", size=11, color="000000")
    font_italic = Font(name="Calibri", size=10, italic=True, color="555555")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    fill_header = PatternFill(start_color=color_primary, end_color=color_primary, fill_type="solid")
    fill_cream = PatternFill(start_color=color_primary_light, end_color=color_primary_light, fill_type="solid")
    fill_accent = PatternFill(start_color=color_accent, end_color=color_accent, fill_type="solid")
    fill_accent_light = PatternFill(start_color=color_accent_light, end_color=color_accent_light, fill_type="solid")
    fill_red = PatternFill(start_color=color_alert, end_color=color_alert, fill_type="solid")
    fill_yellow = PatternFill(start_color=color_warning, end_color=color_warning, fill_type="solid")
    fill_green = PatternFill(start_color=color_ok, end_color=color_ok, fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color=color_gray_border)
    double_border_side = Side(border_style="double", color="000000")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    double_bottom_border = Border(top=thin_border_side, bottom=double_border_side)
    
    # -------------------------------------------------------------
    # SHEET 1: SINTESI EXECUTIVE
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Sintesi Executive"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws1.merge_cells("A1:G2")
    title_cell = ws1["A1"]
    title_cell.value = f"REPORT DI OTTIMIZZAZIONE LOGISTICA E COSTI UPS"
    title_cell.font = font_title
    title_cell.fill = fill_header
    title_cell.alignment = align_center
    
    # Subtitle
    ws1.merge_cells("A3:G3")
    subtitle_cell = ws1["A3"]
    subtitle_cell.value = f"Analisi Codice Cliente: {client_code}  |  Elaborato in automatico sulla base della Guida dei Servizi UPS"
    subtitle_cell.font = font_italic
    subtitle_cell.alignment = align_center
    
    # Spesa Totale and KPIs Section
    ws1["A5"] = "SINTESI FATTURAZIONE"
    ws1["A5"].font = font_section
    
    headers_summary = ["Voce di Spesa", "Valore", "Incidenza su Netto", "Note / Definizione"]
    for col_idx, h in enumerate(headers_summary, start=1):
        cell = ws1.cell(row=6, column=col_idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_left if col_idx != 2 and col_idx != 3 else align_right
        cell.border = thin_border
        
    totale_fattura = grouped["Totale_Spedizione"].sum()
    totale_nolo = grouped["Nolo"].sum()
    totale_fuel = grouped["Fuel"].sum()
    totale_tax = grouped["Tax"].sum()
    totale_supplementi = grouped["Supplementi"].sum()
    
    # Net costs
    spesa_netta = totale_fattura - totale_tax
    totale_anomali = df[df['Alert_Final'] == True]['Importo netto'].sum()
    
    summary_data = [
        ("Nolo Base Spedizioni", totale_nolo, (totale_nolo / spesa_netta) if spesa_netta > 0 else 0, "Tariffa di trasporto base contrattualizzata"),
        ("Fuel Surcharge (Carburante)", totale_fuel, (totale_fuel / spesa_netta) if spesa_netta > 0 else 0, "Supplemento carburante indicizzato mensilmente"),
        ("Supplementi Operativi (Surcharges)", totale_supplementi, (totale_supplementi / spesa_netta) if spesa_netta > 0 else 0, "Spese per servizi accessori e anomalie logistiche"),
        ("Spesa Netta (Senza Tasse)", spesa_netta, 1.0, "Totale spese al netto di IVA ed imposte"),
        ("Tasse / Imposte / IVA", totale_tax, 0, "Imposte e tasse doganali/IVA"),
        ("SPESA TOTALE FATTURA", totale_fattura, 0, "Totale complessivo della fattura UPS analizzata"),
    ]
    
    for row_idx, (voce, valore, incidenza, note) in enumerate(summary_data, start=7):
        r_voce = ws1.cell(row=row_idx, column=1, value=voce)
        r_valore = ws1.cell(row=row_idx, column=2, value=valore)
        r_inc = ws1.cell(row=row_idx, column=3, value=incidenza)
        r_note = ws1.cell(row=row_idx, column=4, value=note)
        
        # Formats
        r_valore.number_format = '€ #,##0.00'
        r_inc.number_format = '0.0%'
        
        # Styles
        r_voce.border = thin_border
        r_valore.border = thin_border
        r_inc.border = thin_border
        r_note.border = thin_border
        
        if voce in ["Spesa Netta (Senza Tasse)", "SPESA TOTALE FATTURA"]:
            r_voce.font = font_bold
            r_valore.font = font_bold
            r_inc.font = font_bold
            r_voce.fill = fill_cream
            r_valore.fill = fill_cream
            r_inc.fill = fill_cream
        else:
            r_voce.font = font_regular
            r_valore.font = font_regular
            r_inc.font = font_regular
            
    # Section 3: Inefficiencies and Opportunities
    ws1["A15"] = "AREE DI EFFICIENTAMENTO E RISPARMIO RECUPERABILE"
    ws1["A15"].font = font_section
    
    headers_opp = ["Opportunità di Risparmio", "Impatto Economico", "Stato", "Raccomandazione Operativa / Azione Correttiva"]
    for col_idx, h in enumerate(headers_opp, start=1):
        cell = ws1.cell(row=16, column=col_idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_left if col_idx != 2 else align_right
        cell.border = thin_border
    
    # 1. Address Correction Impact
    desc_col = 'Descrizione spesa'
    importo_col = 'Importo netto'
    df_evitabili = df[df['Alert_Final'] == True].copy()
    
    costo_indirizzo = 0.0
    num_indirizzi = 0
    if not df_evitabili.empty:
        df_indirizzo = df_evitabili[df_evitabili[desc_col].astype(str).str.lower().str.contains('address|indirizz')]
        costo_indirizzo = df_indirizzo[importo_col].sum()
        num_indirizzi = len(df_indirizzo)
        
    rec_indirizzo = (
        f"Rilevati {num_indirizzi} errori di indirizzo. Si raccomanda di integrare un validatore di CAP/indirizzi "
        "al checkout dell'e-commerce o formare lo staff per verificare le anagrafiche dei clienti prima dell'invio."
    ) if num_indirizzi > 0 else "Nessuna penale rilevata per indirizzi errati. Processo ottimale."
    
    # 2. Volumetric Weight Impact (Aria Pagata)
    df_vol = grouped[(grouped['Peso_Fatt'] > grouped['Peso_Spec']) & (grouped['Peso_Spec'] > 0)]
    num_vol = len(df_vol)
    costo_esposto_vol = df_vol['Totale_Spedizione'].sum()
    pct_vol = (num_vol / len(grouped)) * 100 if len(grouped) > 0 else 0
    
    rec_vol = (
        f"Il {pct_vol:.1f}% delle spedizioni ({num_vol} colli) è stato tassato sul peso volumetrico anziché reale. "
        "Si consiglia di ottimizzare le dimensioni delle scatole, evitare spazi vuoti ed eliminare scatole standard sovradimensionate."
    ) if num_vol > 0 else "Tutti i colli sono stati tassati su peso reale. Ottimo confezionamento."
    
    # 3. Correction Fee Risk
    df_rischio = grouped[grouped['Rischio_SCC'] == True]
    num_rischio = len(df_rischio)
    # Penale stimata
    tot_scc = df_rischio['Totale_Spedizione'].apply(lambda x: max(rules["correction_fee_min"], x * rules["correction_fee_pct"])).sum()
    
    rec_scc = (
        f"Rilevato scostamento peso ≥ {rules['correction_fee_threshold']}% su {num_rischio} colli. Rischio penale Correction Fee. "
        "Si raccomanda di calibrare regolarmente le bilance del magazzino e trasmettere a UPS i pesi reali esatti."
    ) if num_rischio > 0 else "I pesi dichiarati coincidono con quelli rilevati da UPS. Nessun rischio penale."
    
    # 4. Evitable Surcharges (Large package / Additional handling)
    costo_handling = 0.0
    costo_pacco_grande = 0.0
    costo_peso_eccessivo = 0.0
    
    if not df_evitabili.empty:
        df_evitabili['desc_lower'] = df_evitabili[desc_col].astype(str).str.lower()
        costo_handling = df_evitabili[df_evitabili['desc_lower'].str.contains('handling|movimentazione')][importo_col].sum()
        costo_pacco_grande = df_evitabili[df_evitabili['desc_lower'].str.contains('large|pacco.*grande')][importo_col].sum()
        costo_peso_eccessivo = df_evitabili[df_evitabili['desc_lower'].str.contains('over.*max|peso.*eccessivo')][importo_col].sum()
        
    costo_extra_evitabile = costo_handling + costo_pacco_grande + costo_peso_eccessivo
    
    rec_extra = []
    if costo_handling > 0:
        rec_extra.append("Movimentazione aggiuntiva (€ {:.2f}): imballare sempre in cartone ondulato standard ed evitare forme cilindriche/film plastici esterni.".format(costo_handling))
    if costo_pacco_grande > 0:
        rec_extra.append("Pacco Grande (€ {:.2f}): progettare scatole con somma lunghezza + perimetro < 300 cm.".format(costo_pacco_grande))
    if costo_peso_eccessivo > 0:
        rec_extra.append("Peso Eccessivo (€ {:.2f}): non superare mai i 70 kg per singolo collo, utilizzare pallet (Freight).".format(costo_peso_eccessivo))
        
    rec_extra_str = " ".join(rec_extra) if rec_extra else "Nessun costo extra rilevato per imballaggi speciali o fuori misura."
    
    # 5. Domicilio Privato (Residenziale)
    costo_residenziale = 0.0
    num_residenziali = 0
    if not df.empty:
        # Check in the whole invoice if it has residential charges
        df_res = df[df[desc_col].astype(str).str.lower().str.contains('residential|residenziale|domicilio privato')]
        costo_residenziale = df_res[importo_col].sum()
        # count shipments with residential charge
        num_residenziali = df_res[COL_SPEDIZIONE].nunique() if COL_SPEDIZIONE in df_res.columns else len(df_res)
        
    rec_res = (
        f"Spesi € {costo_residenziale:.2f} per consegne a domicili privati ({num_residenziali} spedizioni). "
        "Valutare l'opzione di consegna ai punti di ritiro UPS Access Point o verificare di fatturare correttamente questo costo all'e-commerce."
    ) if costo_residenziale > 0 else "Nessuna consegna residenziale o costo residenziale addebitato."
    
    opp_data = [
        ("Correzione Indirizzi", costo_indirizzo, "⚠️ Critico" if costo_indirizzo > 0 else "✅ Ottimale", rec_indirizzo),
        ("Peso Volumetrico (Aria Pagata)", costo_esposto_vol, "⚠️ Alert" if costo_esposto_vol > 0 else "✅ Ottimale", rec_vol),
        ("Previsione Penali Correction Fee", tot_scc, "🚨 Rischio Elevato" if tot_scc > 0 else "✅ Ottimale", rec_scc),
        ("Supplementi Imballo e Dimensioni", costo_extra_evitabile, "⚠️ Alert" if costo_extra_evitabile > 0 else "✅ Ottimale", rec_extra_str),
        ("Consegne Residenziali (Domicilio)", costo_residenziale, "💡 Info" if costo_residenziale > 0 else "✅ Ottimale", rec_res),
    ]
    
    for row_idx, (opp, impatto, stato, raccomandazione) in enumerate(opp_data, start=17):
        r_opp = ws1.cell(row=row_idx, column=1, value=opp)
        r_imp = ws1.cell(row=row_idx, column=2, value=impatto)
        r_stat = ws1.cell(row=row_idx, column=3, value=stato)
        r_rec = ws1.cell(row=row_idx, column=4, value=raccomandazione)
        
        # Formats
        r_imp.number_format = '€ #,##0.00'
        
        # Styles
        r_opp.font = font_bold
        r_imp.font = font_bold
        r_stat.font = font_header
        r_stat.alignment = align_center
        r_rec.font = font_regular
        
        r_opp.border = thin_border
        r_imp.border = thin_border
        r_stat.border = thin_border
        r_rec.border = thin_border
        
        # Stato color fills
        if "Critico" in stato or "Rischio" in stato:
            r_stat.fill = fill_red
            r_stat.font = Font(name="Calibri", size=11, bold=True, color="900C3F")
        elif "Alert" in stato:
            r_stat.fill = fill_yellow
            r_stat.font = Font(name="Calibri", size=11, bold=True, color="7D6608")
        elif "Info" in stato:
            r_stat.fill = fill_accent_light
            r_stat.font = Font(name="Calibri", size=11, bold=True, color="34495E")
        else:
            r_stat.fill = fill_green
            r_stat.font = Font(name="Calibri", size=11, bold=True, color="196F3D")
            
    # Auto-adjust columns width for Sheet 1
    for col in ws1.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter in ['A', 'B', 'C']:
            # Set fixed appropriate widths
            ws1.column_dimensions[col_letter].width = 25 if col_letter != 'B' else 20
        elif col_letter == 'D':
            ws1.column_dimensions[col_letter].width = 95
        else:
            ws1.column_dimensions[col_letter].width = 15
            
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 100
    
    # -------------------------------------------------------------
    # SHEET 2: DETTAGLIO ANALISI SPEDIZIONI
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Spedizioni Critiche")
    ws2.views.sheetView[0].showGridLines = True
    
    headers_detail = [
        "Numero Spedizione", "Stato Analisi", "Servizio", "Colli", 
        "Peso Dichiarato (Kg)", "Peso Fatturato UPS (Kg)", "Scostamento Peso", 
        "Spesa Totale (€)", "Nolo (€)", "Fuel (€)", "Tax (€)", "Supplementi (€)", "Dettaglio Anomalie Rilevate"
    ]
    
    for col_idx, h in enumerate(headers_detail, start=1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx in [1, 2, 4, 7] else align_left
        cell.border = thin_border
        
    # Sort shipments by total cost descending to highlight the most expensive ones
    grouped_sorted = grouped.sort_values(by="Totale_Spedizione", ascending=False)
    
    row_count = 2
    for _, row in grouped_sorted.iterrows():
        # Filter for criticità: alert present OR volumetric weight issue OR correction fee risk
        has_alert = row.get("Presenza_Alert", False)
        peso_fatt = row.get("Peso_Fatt", 0.0)
        peso_spec = row.get("Peso_Spec", 0.0)
        has_vol = (peso_fatt > peso_spec) and (peso_spec > 0)
        has_scc = row.get("Rischio_SCC", False)
        
        # If it has no issues, we don't show it or show it with normal style. 
        # To keep the report focused on optimization, let's show ALL but highlight the critical ones.
        # This gives a complete reference.
        
        status_lbl = "⚠️ Alert" if has_alert else "✅ OK"
        if has_scc:
            status_lbl = "🚨 Rischio Correction"
            
        anomalie = []
        if has_scc:
            anomalie.append(f"Scostamento Peso ≥ {rules['correction_fee_threshold']}%")
        if has_vol:
            anomalie.append("Peso Volumetrico > Peso Reale")
        if has_alert:
            anomalie.append("Presenza di Supplemento Anomalo")
            
        # Get actual descriptions of alert surcharges for this tracking
        trk_details = df[(df[COL_SPEDIZIONE] == row[COL_SPEDIZIONE]) & (df['Alert_Final'] == True)]
        for _, det in trk_details.iterrows():
            anomalie.append(str(det[COL_DESCRIZIONE]))
            
        anomalie_str = " | ".join(list(set(anomalie))) if anomalie else "Nessuna"
        
        # Delta peso percent
        delta_p = ((peso_fatt - peso_spec) / peso_spec) if peso_spec > 0 else 0.0
        
        r_data = [
            row[COL_SPEDIZIONE],
            status_lbl,
            row.get("Servizio", "Sconosciuto"),
            row.get("Pacchi_Display", "1 collo"),
            peso_spec,
            peso_fatt,
            delta_p if delta_p > 0 else 0,
            row.get("Totale_Spedizione", 0.0),
            row.get("Nolo", 0.0),
            row.get("Fuel", 0.0),
            row.get("Tax", 0.0),
            row.get("Supplementi", 0.0),
            anomalie_str
        ]
        
        for col_idx, val in enumerate(r_data, start=1):
            cell = ws2.cell(row=row_count, column=col_idx, value=val)
            cell.border = thin_border
            cell.font = font_regular
            
            # Format types
            if col_idx in [5, 6]:
                cell.number_format = '#,##0.00'
                cell.alignment = align_right
            elif col_idx == 7:
                cell.number_format = '0.0%'
                cell.alignment = align_right
            elif col_idx in [8, 9, 10, 11, 12]:
                cell.number_format = '€ #,##0.00'
                cell.alignment = align_right
            elif col_idx in [1, 2, 4]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            # Formatting highlights
            if col_idx == 2:
                if "Rischio" in val:
                    cell.fill = fill_red
                    cell.font = Font(name="Calibri", size=11, bold=True, color="900C3F")
                elif "Alert" in val:
                    cell.fill = fill_yellow
                    cell.font = Font(name="Calibri", size=11, bold=True, color="7D6608")
                else:
                    cell.fill = fill_green
                    cell.font = Font(name="Calibri", size=11, bold=True, color="196F3D")
                    
            # Soft red fill on rows with Correction Fee risk
            if has_scc and col_idx != 2:
                cell.fill = PatternFill(start_color="FFF2F2", end_color="FFF2F2", fill_type="solid")
            # Soft yellow fill on rows with simple alerts
            elif has_alert and not has_scc and col_idx != 2:
                cell.fill = PatternFill(start_color="FFFFE6", end_color="FFFFE6", fill_type="solid")
                
        row_count += 1
        
    # Auto-adjust column widths for Sheet 2
    for col in ws2.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)
    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['C'].width = 24
    ws2.column_dimensions['M'].width = 50
    
    # Save Workbook to memory stream
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream.getvalue()
